import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import Workbook

# 1. BASE DE DATOS INTEGRADA (Extraída de tus archivos subidos)
data = {
    'Equipo': [
        'Agitador Magnetico (16 posiciones)', 'Balanza (max 100kg)', 'Horno de secado', 
        'Campana de extracción', 'Balanza Analitica', 'Medidor NO', 'Balanza digital (30kg)',
        'Anemómetro', 'Anemómetro', 'Multiparametro portatil', 'Alzador electrico', 
        'Hidrolavadora GHP200', 'Balanza Digital (30Kg)', 'Anemometro (198)'
    ],
    'Codigo_Sistema': [
        'EQ-CB-023', 'EQ-CB-217', 'EQ-CB-066', 'EQ-CB-092', 'EQ-CB-021', 
        'EQ-LIX-010', 'EQ-CB-243', 'EQ-CB-254', 'EQ-CB-255', 'SIN CODIGO', 
        'SIN CODIGO', 'EQ-CB-258', 'EQ-CB-124', 'EQ-CB-198'
    ],
    'Observacion_Auditoria': [
        'Código coincide', 'Código coincide', 'Código coincide', 'Código coincide', 
        'EQUIPO DE BAJA', 'ETIQUETADO COMO EQ-CB-152 (ERROR)', 'CÓDIGO REPETIDO', 
        'CÓDIGO REPETIDO', 'CÓDIGO REPETIDO', 'FALTA CODIFICAR (REF EQ-CB-221)', 
        'FALTA CODIFICAR', 'Sin fecha de mantención', 'EQUIPO DE BAJA', 'Sin fecha de mantención'
    ]
}

df = pd.DataFrame(data)

# 2. ASIGNACIÓN DE ESTADOS PARA EL REPORTE
def definir_estado_final(row):
    obs = str(row['Observacion_Auditoria']).upper()
    cod = str(row['Codigo_Sistema']).upper()
    
    if 'BAJA' in obs: return 'DE BAJA (AZUL)'
    if 'SIN CODIGO' in cod or 'FALTA CODIFICAR' in obs: return 'SIN CÓDIGO (NARANJO)'
    if 'ETIQUETADO COMO' in obs or 'ERROR' in obs: return 'CÓDIGO NO COINCIDE (VERDE)'
    if 'REPETIDO' in obs: return 'CÓDIGO REPETIDO (VERDE AZULADO)'
    return 'COINCIDE (ROSADO)'

df['Estado Final'] = df.apply(definir_estado_final, axis=1)

# 3. GENERACIÓN DEL EXCEL CON COLORES ARREGLADOS (aRGB)
nombre_archivo = 'Reporte_Final_TechOps_Arreglado.xlsx'
writer = pd.ExcelWriter(nombre_archivo, engine='openpyxl')
df.to_excel(writer, index=False, sheet_name='Auditoria Tech Ops')

ws = writer.sheets['Auditoria Tech Ops']

# Prefijo 'FF' añadido para evitar el error ValueError: Colors must be aRGB
colores_argb = {
    'COINCIDE (ROSADO)': 'FFFFC0CB',             # Rosa opaco
    'CÓDIGO REPETIDO (VERDE AZULADO)': 'FF80CBC4',   # Verde azulado opaco
    'CÓDIGO NO COINCIDE (VERDE)': 'FFC8E6C9',       # Verde claro opaco
    'SIN CÓDIGO (NARANJO)': 'FFFFE0B2',             # Naranjo claro opaco
    'DE BAJA (AZUL)': 'FFBBDEFB'                    # Azul claro opaco
}

# Aplicar el relleno de color según el estado
for row in range(2, len(df) + 2):
    estado = ws.cell(row=row, column=4).value # La columna 4 es 'Estado Final'
    if estado in colores_argb:
        color_hex = colores_argb[estado]
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
        ws.cell(row=row, column=4).fill = fill

writer.close()
print(f"✅ ¡Éxito! Reporte generado como: {nombre_archivo}")
